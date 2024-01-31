import sys, os
sys.path.append(os.path.abspath("./src"))
import openpyxl
import matplotlib.pyplot as plt
import chart_parametrs

DIR_PATH = os.path.abspath("example")

if __name__ == "__main__":
    wb_obj = openpyxl.load_workbook(DIR_PATH + "/data/example.xlsx")
    sheet_obj = wb_obj.active
    max_row = sheet_obj.max_row
    deformations = []
    tensions = []
    for i in range(1, max_row + 1):
        deformations.append(sheet_obj.cell(row = i, column = 1).value)
        tensions.append(sheet_obj.cell(row = i, column = 2).value)
    
    diagramm = chart_parametrs.ChartParametrs(deformations, tensions)
    diagramm.fit()
    chart_parametrs.make_plot(diagramm).savefig(DIR_PATH + "/img/example.pdf")